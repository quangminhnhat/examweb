#!/usr/bin/env python3
"""
Exam Import XLSX Generator
Creates a properly formatted Excel file for importing exams into the system.
"""

import openpyxl
from openpyxl.styles import Font, Alignment
import sys
import os

def create_exam_import_xlsx(filename="exam_import_template.xlsx"):
    """Create an XLSX file with exam import template structure."""

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Import"

    # Define styles
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    center_align = Alignment(horizontal='center', vertical='center')

    # Row 0: Exam metadata
    ws['A1'] = "Computer Science Final Exam"  # Exam title
    ws['B1'] = "90"  # Duration in minutes
    ws['C1'] = "150"  # Total score

    # Style the metadata row
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}1']
        cell.font = header_font
        cell.alignment = center_align

    # Row 2: Headers (will be skipped during import)
    headers = ["Question Content", "Option A", "Option B", "Option C", "Option D", "Correct Option", "Score"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.alignment = center_align

    # Sample questions data
    questions = [
        {
            "content": "What does CPU stand for?",
            "options": ["Central Processing Unit", "Central Program Unit", "Computer Personal Unit", "Central Processor Unit"],
            "correct": "A",
            "score": 2
        },
        {
            "content": "Which programming language is known as the 'mother of all languages'?",
            "options": ["C", "C++", "Java", "Python"],
            "correct": "A",
            "score": 2
        },
        {
            "content": "What is the time complexity of binary search?",
            "options": ["O(n)", "O(log n)", "O(n^2)", "O(1)"],
            "correct": "B",
            "score": 3
        },
        {
            "content": "Which data structure uses LIFO (Last In, First Out)?",
            "options": ["Queue", "Stack", "Array", "Linked List"],
            "correct": "B",
            "score": 2
        },
        {
            "content": "What does HTML stand for?",
            "options": ["HyperText Markup Language", "High Tech Modern Language", "Home Tool Markup Language", "Hyper Transfer Markup Language"],
            "correct": "A",
            "score": 2
        },
        {
            "content": "Which of these is not a programming paradigm?",
            "options": ["Object-oriented", "Functional", "Procedural", "Algorithmic"],
            "correct": "D",
            "score": 2
        },
        {
            "content": "What is the purpose of a constructor in OOP?",
            "options": ["To destroy objects", "To initialize objects", "To copy objects", "To compare objects"],
            "correct": "B",
            "score": 3
        },
        {
            "content": "Which HTTP status code indicates 'Not Found'?",
            "options": ["200", "301", "404", "500"],
            "correct": "C",
            "score": 2
        },
        {
            "content": "What does SQL stand for?",
            "options": ["Simple Query Language", "Structured Query Language", "Standard Question Language", "System Query Language"],
            "correct": "B",
            "score": 2
        },
        {
            "content": "Which sorting algorithm has the best average case time complexity?",
            "options": ["Bubble Sort", "Quick Sort", "Merge Sort", "Insertion Sort"],
            "correct": "C",
            "score": 3
        }
    ]

    # Add questions starting from row 3
    for row_num, question in enumerate(questions, 3):
        # Question content
        ws.cell(row=row_num, column=1).value = question["content"]

        # Options A-D
        for col_num, option in enumerate(question["options"], 2):
            ws.cell(row=row_num, column=col_num).value = option

        # Correct option
        ws.cell(row=row_num, column=6).value = question["correct"]

        # Score
        ws.cell(row=row_num, column=7).value = question["score"]

        # Style the row
        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = normal_font
            if col_num == 6:  # Correct option column
                cell.alignment = center_align

    # Auto-adjust column widths
    for col_num in range(1, 8):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, len(questions) + 3):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)  # Max width of 50

    # Save the workbook
    wb.save(filename)
    print(f"✅ XLSX file created: {filename}")
    print(f"📊 Exam: Computer Science Final Exam")
    print(f"⏱️  Duration: 90 minutes")
    print(f"🎯 Total Score: 150 points")
    print(f"❓ Questions: {len(questions)}")
    print("\n📋 File Structure:")
    print("   Row 1: Exam metadata (Title, Duration, Total Score)")
    print("   Row 2: Headers (skipped during import)")
    print("   Row 3+: Questions with options and correct answers")

def create_simple_template(filename="simple_exam_template.xlsx"):
    """Create a simpler template with just a few questions."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simple Exam"

    # Exam metadata
    ws['A1'] = "Math Quiz"
    ws['B1'] = "30"
    ws['C1'] = "50"

    # Headers
    headers = ["Question Content", "Option A", "Option B", "Option C", "Option D", "Correct Option", "Score"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=2, column=col_num).value = header

    # Simple questions
    simple_questions = [
        ["What is 2 + 2?", "3", "4", "5", "6", "B", "1"],
        ["What is 10 - 5?", "3", "5", "7", "9", "B", "1"],
        ["What is 3 * 4?", "7", "12", "15", "18", "B", "2"],
        ["What is 20 / 4?", "3", "4", "5", "6", "C", "2"],
        ["What is 15 + 7?", "20", "21", "22", "23", "C", "1"]
    ]

    for row_num, question in enumerate(simple_questions, 3):
        for col_num, value in enumerate(question, 1):
            ws.cell(row=row_num, column=col_num).value = value

    wb.save(filename)
    print(f"✅ Simple template created: {filename}")

if __name__ == "__main__":
    print("🎓 Exam Import XLSX Generator")
    print("=" * 40)

    # Check if openpyxl is available
    try:
        import openpyxl
    except ImportError:
        print("❌ Error: openpyxl library is required.")
        print("   Install it with: pip install openpyxl")
        sys.exit(1)

    # Create the main template
    create_exam_import_xlsx()

    # Create a simple template too
    create_simple_template()

    print("\n🚀 You can now use these XLSX files to import exams into your system!")
    print("   Upload them through the exam import feature in your application.")